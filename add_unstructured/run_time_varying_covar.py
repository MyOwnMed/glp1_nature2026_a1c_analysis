#!/usr/bin/env python3
"""Time-varying weight covariate sensitivity analysis.

Fits GEE models controlling for concurrent weight change (pct_weight_change)
in addition to demographics, to assess whether outcome trajectories are
mediated by or independent of weight loss after GLP-1 initiation.

For each domain (PHQ-9, Pain, Alcohol, Waist, Muscle):
  1. Load domain prepared CSV.
  2. Merge concurrent weight data from step1 CSV by nearest date (within ±60 d).
  3. Fit two GEE B-spline models:
       - Base:   outcome ~ bs(days) + demographics
       - Weight: outcome ~ bs(days) + pct_weight_change + demographics
  4. Report the pct_weight_change coefficient and compare temporal trajectories.
  5. Also fit CFB versions: change_score ~ bs(days) + [pct_weight_change] + demographics.

Output → <AU_OUTROOT>/time_varying_covar/figures/ and /tables/

Environment variables:
  AU_DATADIR  — domain CSV directory (required)
  AU_OUTROOT  — output root (required)
  AU_STEP1    — step1 analysis_ready CSV with weight data (required)
"""

import logging, os, warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from pathlib import Path
from patsy import dmatrices, build_design_matrices
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Gaussian
from statsmodels.genmod.cov_struct import Independence

# Confidence multiplier derived from a single configurable confidence level
# (code-review item 9); replaces a hard-coded 1.96. See analysis_config.py.
import sys as _sys
from pathlib import Path as _Path

for _p in [_Path(__file__).resolve().parent, *_Path(__file__).resolve().parents]:
    if (_p / "analysis_config.py").exists():
        _sys.path.insert(0, str(_p))
        break
from analysis_config import z_critical

Z_CRIT = z_critical()

warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT  = Path(__file__).resolve().parents[2]
DATA  = Path(os.environ.get('AU_DATADIR', str(ROOT / "output" / "submitted_analysis" / "1_no_adherence" / "data")))
_AU   = os.environ.get('AU_OUTROOT')
OUT   = (Path(_AU) / "time_var_covar") if _AU \
        else ROOT / "output" / "submitted_analysis" / "1_no_adherence" / "time_var_covar"
FIGS  = OUT / "figures"
TABS  = OUT / "tables"
for d in [FIGS, TABS]:
    d.mkdir(parents=True, exist_ok=True)

# Step1 CSV with weight data (pct_weight_change, weight_in_pounds_final)
_STEP1 = os.environ.get('AU_STEP1')
STEP1_PATH = Path(_STEP1) if _STEP1 else (
    ROOT / "output" / "step1_prepare_analysis_dataset" / "analysis_ready_gap120.csv")

# Full merged file — used to supplement step1 with weight data for patients
# who have no A1c and were therefore excluded from the step1 A1c+weight cohort.
STEP8G_PATH = ROOT / "root_data" / "merged" / \
    "step8g_with_unstructured_flags_with_assessments_weightcleaned.csv"

DOMAINS = {
    'phq9': {
        'label':            'Depression (PHQ-9, 0–27 scale)',
        'ylabel':           'PHQ-9 Score (0–27)',
        'val':              'phq9_value',
        'elev_threshold':   5,
        'elev_label':       'Elevated baseline ≥5 (mild+)',
        'elev_threshold_2': 10,
        'elev_label_2':     'Elevated baseline ≥10 (moderate-severe)',
        'spline_df':        3,
        'color_all':        '#7BAFD4',
        'color_elev':       '#1A5C8A',
        'color_elev_2':     '#0A2340',
        'extra_cov':        ['antidepressant_baseline', 'covid_era'],
    },
    'pain_score': {
        'label':            'General Pain Intensity (0–10 scale)',
        'ylabel':           'Pain Score (0–10)',
        'val':              'pain_score_value',
        'elev_threshold':   4,
        'elev_label':       'Elevated baseline ≥4 (moderate+)',
        'elev_threshold_2': 7,
        'elev_label_2':     'Elevated baseline ≥7 (severe)',
        'spline_df':        3,
        'color_all':        '#F5A673',
        'color_elev':       '#C0392B',
        'color_elev_2':     '#7B0000',
        'extra_cov':        ['covid_era'],
    },
    'waist_circumference': {
        'label':            'Waist Circumference (Inches)',
        'ylabel':           'Waist Circumference (inches)',
        'val':              'waist_circumference_value',
        'elev_threshold':   None,
        'elev_label':       None,
        'elev_threshold_2': None,
        'elev_label_2':     None,
        'spline_df':        3,
        'color_all':        '#5DAD6F',
        'color_elev':       None,
        'color_elev_2':     None,
        'extra_cov':        [],
    },
    'alcohol': {
        'label':            'Alcohol Use (Drinks/day)',
        'ylabel':           'Alcohol Score',
        'val':              'alcohol_value',
        'elev_threshold':   'top25pct',
        'elev_label':       'Top 25% baseline drinkers',
        'elev_threshold_2': None,
        'elev_label_2':     None,
        'spline_df':        3,
        'color_all':        '#9B59B6',
        'color_elev':       '#6C1A8A',
        'color_elev_2':     None,
        'extra_cov':        ['covid_era'],
    },
    'muscle_strength': {
        'label':            'Muscle Strength (MRC, 0–5 scale)',
        'ylabel':           'Strength Score',
        'val':              'muscle_strength_value',
        'elev_threshold':   None,
        'elev_label':       None,
        'elev_threshold_2': None,
        'elev_label_2':     None,
        'spline_df':        3,
        'color_all':        '#95A5A6',
        'color_elev':       None,
        'color_elev_2':     None,
        'extra_cov':        ['covid_era'],
    },
}

plt.rcParams.update({
    'font.size': 11, 'axes.titlesize': 13, 'axes.labelsize': 12,
    'figure.dpi': 150, 'savefig.dpi': 300, 'savefig.bbox': 'tight',
    'font.family': 'sans-serif',
    'axes.spines.top': False, 'axes.spines.right': False,
})

BASELINE_WINDOW = 30
FOLLOWUP_START  = 30
FOLLOWUP_END    = 365
DAYS_GRID       = np.arange(-180, 366, 7)
DAYS_GRID_CFB   = np.arange(0, FOLLOWUP_END + 1, 7)
TIMEPOINTS      = [(91, '3 mo'), (183, '6 mo'), (274, '9 mo'), (365, '12 mo')]

BASE_COV_COLS = [
    'age', 'gender', 'race',
    'baseline_a1c_category', 'baseline_bmi_final_category',
]
BASE_COV_FORMULA = (
    "age + C(gender) + C(race) + "
    "C(baseline_a1c_category) + C(baseline_bmi_final_category)"
)


# ── Weight data loading ────────────────────────────────────────────────────

def load_weight_series():
    """Load per-patient time-varying weight from step1 CSV.

    step1 only includes patients who have both weight AND A1c data (it was
    built for the primary A1c analysis).  The domain analyses (PHQ-9, pain,
    etc.) include a broader population — patients with domain observations but
    no A1c.  For those patients we supplement step1 with weight measurements
    extracted directly from the full merged step8g file, so that the TVC
    analysis covers the same population as the Forest/CFB analyses.
    """
    frames = []

    # ── 1. Primary source: step1 (fast, pre-computed) ────────────────────
    step1_pts: set = set()
    if STEP1_PATH.exists():
        try:
            s1 = pd.read_csv(STEP1_PATH,
                             usecols=['patient_id', 'days_from_baseline',
                                      'pct_weight_change'])
            s1['days_from_baseline'] = pd.to_numeric(
                s1['days_from_baseline'], errors='coerce')
            s1['pct_weight_change']  = pd.to_numeric(
                s1['pct_weight_change'],  errors='coerce')
            s1 = s1.dropna(subset=['days_from_baseline', 'pct_weight_change'])
            step1_pts = set(s1['patient_id'].unique())
            frames.append(s1[['patient_id', 'days_from_baseline',
                               'pct_weight_change']])
            log.info("step1 weight: %d rows, %d patients",
                     len(s1), len(step1_pts))
        except Exception as e:
            log.warning("Could not load step1 weight data: %s", e)
    else:
        log.warning("Step1 file not found: %s", STEP1_PATH)

    # ── 2. Supplemental source: step8g for patients not in step1 ─────────
    # These are domain patients who have weight measurements in the EHR but
    # were excluded from the A1c+weight primary analysis (e.g. no A1c).
    if STEP8G_PATH.exists():
        try:
            log.info("Loading supplemental weight from step8g …")
            s8 = pd.read_csv(
                STEP8G_PATH,
                usecols=['patient_id', 'days_from_baseline',
                         'weight_in_pounds_final', 'baseline_weight_final'],
            )
            # Keep only rows with an actual weight measurement
            s8 = s8[s8['weight_in_pounds_final'].notna() &
                    s8['baseline_weight_final'].notna()].copy()
            # Exclude patients already covered by step1
            if step1_pts:
                s8 = s8[~s8['patient_id'].isin(step1_pts)]
            s8['days_from_baseline'] = pd.to_numeric(
                s8['days_from_baseline'], errors='coerce')
            s8['weight_in_pounds_final']  = pd.to_numeric(
                s8['weight_in_pounds_final'], errors='coerce')
            s8['baseline_weight_final']   = pd.to_numeric(
                s8['baseline_weight_final'],  errors='coerce')
            s8 = s8.dropna(subset=['days_from_baseline',
                                    'weight_in_pounds_final',
                                    'baseline_weight_final'])
            # Compute per-row pct_weight_change
            s8['pct_weight_change'] = (
                100.0 * (s8['weight_in_pounds_final'] - s8['baseline_weight_final'])
                / s8['baseline_weight_final']
            )
            s8 = s8.dropna(subset=['pct_weight_change'])
            log.info("step8g supplemental weight: %d rows, %d patients",
                     len(s8), s8['patient_id'].nunique())
            frames.append(s8[['patient_id', 'days_from_baseline',
                               'pct_weight_change']])
        except Exception as e:
            log.warning("Could not load step8g supplemental weight: %s", e)
    else:
        log.warning("step8g not found at %s — no supplemental weight", STEP8G_PATH)

    if not frames:
        log.error("No weight data loaded — cannot run time-varying analysis.")
        return pd.DataFrame()

    df = pd.concat(frames, ignore_index=True)
    df = df.sort_values(['patient_id', 'days_from_baseline'])
    log.info("Combined weight series: %d rows, %d patients",
             len(df), df['patient_id'].nunique())
    return df[['patient_id', 'days_from_baseline', 'pct_weight_change']]


def load_domain(domain, meta):
    """Load domain CSV (same population as CFB analysis — no has_both_periods filter)."""
    p = DATA / f"{domain}_prepared.csv"
    if not p.exists():
        return pd.DataFrame()
    df = pd.read_csv(p)
    df['days_from_baseline'] = pd.to_numeric(df['days_from_baseline'], errors='coerce')
    val = meta['val']
    df = df.dropna(subset=[val, 'days_from_baseline'])
    return df


def filter_elevated(df, val_col, threshold):
    """Filter to patients whose baseline anchor score (±30d) >= threshold.

    threshold may be a numeric value, None (skip), or 'top25pct' to compute
    the 75th percentile of baseline scores dynamically.
    """
    if threshold is None or df.empty:
        return None
    base  = df[df['days_from_baseline'].between(-BASELINE_WINDOW, BASELINE_WINDOW)]
    means = base.groupby('patient_id')[val_col].mean()
    if threshold == 'top25pct':
        threshold = float(means.quantile(0.75))
        log.info("    top25pct threshold for %s: %.2f", val_col, threshold)
    pids  = set(means[means >= threshold].index)
    return df[df.patient_id.isin(pids)].copy() if pids else None


def merge_weight(domain_df: pd.DataFrame, weight_df: pd.DataFrame) -> pd.DataFrame:
    """Assign the most recent (LOCF) weight observation to each domain assessment row.

    Uses pd.merge_asof with direction='nearest' so that every patient who has
    ANY weight measurement (at minimum their baseline weight) receives a
    pct_weight_change value.  No observations are dropped for lacking a
    concurrent weight — patients with only a baseline weight will carry that
    value forward across all follow-up rows, matching the population used by
    the Forest / CFB analyses.
    """
    if weight_df.empty:
        domain_df = domain_df.copy()
        domain_df['pct_weight_change'] = np.nan
        return domain_df

    domain_df = domain_df.copy()
    domain_df['days_from_baseline'] = domain_df['days_from_baseline'].astype(float)
    wt = weight_df[['patient_id', 'days_from_baseline', 'pct_weight_change']].copy()
    wt['days_from_baseline'] = wt['days_from_baseline'].astype(float)

    # merge_asof requires both dataframes sorted by the merge key globally
    dom_sorted = domain_df.sort_values('days_from_baseline')
    wt_sorted  = wt.sort_values('days_from_baseline')

    # LOCF: nearest prior weight; fall back to nearest future if no prior exists
    merged = pd.merge_asof(
        dom_sorted,
        wt_sorted,
        on='days_from_baseline',
        by='patient_id',
        direction='nearest',
    )

    n_matched = merged['pct_weight_change'].notna().sum()
    log.info("    Weight merge (LOCF/nearest): %d / %d rows matched",
             n_matched, len(merged))
    return merged


# ── GEE helpers ────────────────────────────────────────────────────────────

def get_ref_covars(df, extra_cov=None):
    """Reference covariate values for prediction."""
    extra_cov = extra_cov or []
    ref = {}
    if 'age' in df.columns:
        ref['age'] = float(df['age'].mean())
    for c in ['gender', 'race', 'baseline_a1c_category', 'baseline_bmi_final_category']:
        if c in df.columns and not df[c].dropna().empty:
            ref[c] = df[c].mode().iloc[0]
    # Weight reference: mean pct_weight_change (0 = no change)
    if 'pct_weight_change' in df.columns:
        ref['pct_weight_change'] = float(df['pct_weight_change'].mean())
    for c in extra_cov:
        if c in df.columns:
            ref[c] = df[c].mode().iloc[0] if df[c].dtype == object else float(df[c].mean())
    return ref


def fit_gee(df, outcome_col, spline_df, include_weight=False, extra_cov=None):
    """Fit GEE B-spline model, optionally including pct_weight_change.

    Returns (result, design_info, d_min, d_max, ref_covars) or Nones on failure.
    """
    extra_cov = extra_cov or []
    sub_cols = [outcome_col, 'days_from_baseline', 'patient_id'] + BASE_COV_COLS + extra_cov
    if include_weight:
        sub_cols.append('pct_weight_change')
    sub_cols = [c for c in sub_cols if c in df.columns]

    sub = df[sub_cols].dropna(subset=[outcome_col, 'days_from_baseline']).copy()
    # With LOCF weight merge all patients have a weight value; only drop rows
    # where weight is genuinely missing (should not happen after LOCF merge).
    if include_weight:
        sub = sub.dropna(subset=['pct_weight_change'])

    # Clean categoricals
    for c in ['gender', 'race', 'baseline_a1c_category', 'baseline_bmi_final_category']:
        if c in sub.columns:
            sub[c] = sub[c].fillna('Unknown').astype(str)
    if 'age' in sub.columns:
        sub['age'] = pd.to_numeric(sub['age'], errors='coerce').fillna(sub['age'].median())

    if sub.patient_id.nunique() < 15 or len(sub) < 30:
        return None, None, None, None, None

    d_min = float(sub.days_from_baseline.min())
    d_max = float(sub.days_from_baseline.max())
    ref = get_ref_covars(sub, extra_cov=extra_cov if include_weight else extra_cov)

    cov_str = BASE_COV_FORMULA
    xtra_terms = [c for c in extra_cov if c in sub.columns]
    if xtra_terms:
        cov_str += ' + ' + ' + '.join(xtra_terms)
    if include_weight and 'pct_weight_change' in sub.columns:
        cov_str += ' + pct_weight_change'

    formula = (f"{outcome_col} ~ bs(days_from_baseline, df={spline_df}, "
               f"include_intercept=False) + {cov_str}")
    try:
        y, X = dmatrices(formula, sub, return_type='dataframe')
    except Exception as e:
        log.error("  dmatrices failed: %s", e)
        return None, None, None, None, None

    ids = sub.loc[y.index, 'patient_id']
    try:
        result = GEE(y, X, groups=ids, family=Gaussian(),
                     cov_struct=Independence()).fit()
        log.info("  GEE (%s weight=%s): %d obs, %d pts, df=%d",
                 outcome_col, include_weight, len(sub),
                 sub.patient_id.nunique(), spline_df)
        return result, X.design_info, d_min, d_max, ref
    except Exception as e:
        log.error("  GEE failed: %s", e)
        return None, None, None, None, None


def predict_gee(result, design_info, days_grid, d_min, d_max, ref_covars=None):
    """Predict GEE mean + 95% CI on a day grid."""
    out = np.full((3, len(days_grid)), np.nan)
    eps  = 0.5
    mask = (days_grid >= d_min) & (days_grid <= d_max - eps)
    if not mask.any():
        return out[0], out[1], out[2]

    pred_df = pd.DataFrame({'days_from_baseline': days_grid[mask]})
    if ref_covars:
        for col, val in ref_covars.items():
            pred_df[col] = val
    try:
        X_pred = np.asarray(build_design_matrices([design_info], pred_df)[0])
        mv     = np.asarray(result.predict(X_pred))
        V      = np.asarray(result.cov_params())
        se     = np.sqrt(np.clip(np.einsum('ij,jk,ik->i', X_pred, V, X_pred), 0, None))
        out[0][mask] = mv
        out[1][mask] = mv - Z_CRIT * se
        out[2][mask] = mv + Z_CRIT * se
    except Exception as e:
        log.error("  predict failed: %s", e)
    return out[0], out[1], out[2]


def support_mask(df, val_col, days_grid, window=30, min_pts=10):
    sub = df[[val_col, 'days_from_baseline', 'patient_id']].dropna()
    return np.array([
        sub[np.abs(sub.days_from_baseline - d) <= window].patient_id.nunique() >= min_pts
        for d in days_grid
    ])


def xfmt(x, _):
    return '0' if x == 0 else (f'+{int(x)}' if x > 0 else str(int(x)))


# ── Point estimate extraction at 3/6/9/12 months ─────────────────────────────

def predict_at_days(result, design_info, days, d_min, d_max, ref_covars=None):
    """Predict GEE mean + SE at a list of specific days."""
    days_arr = np.array(days, dtype=float)
    mean_v = np.full(len(days_arr), np.nan)
    se_v   = np.full(len(days_arr), np.nan)
    mask = (days_arr >= d_min - 3) & (days_arr <= d_max + 3)
    if not mask.any():
        return mean_v, se_v
    clamped = np.clip(days_arr[mask], d_min, d_max)
    pred_df = pd.DataFrame({'days_from_baseline': clamped})
    if ref_covars:
        for col, val in ref_covars.items():
            pred_df[col] = val
    try:
        X_pred = np.asarray(build_design_matrices([design_info], pred_df)[0])
        mv     = np.asarray(result.predict(X_pred))
        V      = np.asarray(result.cov_params())
        var    = np.clip(np.einsum('ij,jk,ik->i', X_pred, V, X_pred), 0, None)
        mean_v[mask] = mv
        se_v[mask]   = np.sqrt(var)
    except Exception as e:
        log.error('  predict_at_days failed: %s', e)
    return mean_v, se_v


def extract_cfb_point_estimates(domain_label, subgroup_results):
    """Extract Base-CFB and Weight-adjusted-CFB point estimates at 3/6/9/12 months."""
    rows = []
    all_days = [0.0] + [float(d) for d, _ in TIMEPOINTS]

    for sg in subgroup_results:
        if not sg.get('cfb_possible', False):
            continue
        lbl = sg['sg_label']

        for analysis, res, di, mn, mx, ref, n_val in [
            ('TVC_Base',  sg.get('res_cb'), sg.get('di_cb'), sg.get('mn_cb'),
             sg.get('mx_cb'), sg.get('ref_cb'), sg.get('n_cfb')),
            ('TVC_WtAdj', sg.get('res_cw'), sg.get('di_cw'), sg.get('mn_cw'),
             sg.get('mx_cw'), sg.get('ref_cw'), sg.get('n_cfb')),
        ]:
            if res is None or di is None:
                continue
            mean_v, se_v = predict_at_days(res, di, all_days, mn, mx, ref_covars=ref)
            y0   = mean_v[0] if not np.isnan(mean_v[0]) else 0.0
            se0  = se_v[0]   if not np.isnan(se_v[0])   else 0.0
            for i, (day, tp_label) in enumerate(TIMEPOINTS):
                j = i + 1
                if np.isnan(mean_v[j]):
                    continue
                delta    = mean_v[j] - y0
                delta_se = np.sqrt(se_v[j]**2 + se0**2)
                rows.append({
                    'analysis':  analysis,
                    'subgroup':  lbl,
                    'timepoint': tp_label,
                    'day':       day,
                    'n':         n_val,
                    'estimate':  delta,
                    'se':        delta_se,
                    'ci_lo':     delta - Z_CRIT * delta_se,
                    'ci_hi':     delta + Z_CRIT * delta_se,
                    'domain':    domain_label,
                })
    return rows


# ── Forest plot: Base CFB vs Weight-adjusted CFB ───────────────────────────────

def plot_forest_tvc(domain, meta, est_df):
    """Forest plot for TVC: Base CFB (left) vs Weight-Adjusted CFB (right)."""
    from matplotlib.lines import Line2D

    # Build subgroup metadata matching the order/colors in DOMAINS
    subgroups = [
        {'name': 'All patients', 'color': meta['color_all'], 'marker': 'o'},
    ]
    if meta.get('elev_label'):
        subgroups.append({'name': meta['elev_label'],
                          'color': meta['color_elev'], 'marker': 's'})
    if meta.get('elev_label_2') and meta.get('elev_threshold_2') is not None:
        subgroups.append({'name': meta['elev_label_2'],
                          'color': meta['color_elev_2'], 'marker': 'D'})

    timepoint_labels = [t for _, t in TIMEPOINTS]
    n_sg  = len(subgroups)

    fig_h = max(4.0, len(timepoint_labels) * (n_sg * 0.42 + 0.65) + 1.5)
    fig, axes = plt.subplots(1, 2, figsize=(14, fig_h), sharey=True)

    y_gap  = 0.42
    tp_gap = 1.35
    y_positions = {}
    y_tick_pos  = {}
    y_current   = 0

    for tp_label in reversed(timepoint_labels):
        sg_ys = []
        for sg in subgroups:
            y_positions[(tp_label, sg['name'])] = y_current
            sg_ys.append(y_current)
            y_current += y_gap
        y_tick_pos[tp_label] = np.mean(sg_ys)
        y_current += tp_gap - y_gap

    for ax, analysis, panel_title in zip(
            axes,
            ['TVC_Base', 'TVC_WtAdj'],
            ['CFB (unadjusted)', 'CFB + weight-adjusted']):
        sub = est_df[est_df.analysis == analysis]

        for _, row in sub.iterrows():
            key = (row.timepoint, row.subgroup)
            if key not in y_positions:
                continue
            yi  = y_positions[key]
            sg_m = next((s for s in subgroups if s['name'] == row.subgroup), None)
            if sg_m is None:
                continue
            ax.errorbar(
                row.estimate, yi,
                xerr=[[row.estimate - row.ci_lo], [row.ci_hi - row.estimate]],
                fmt=sg_m['marker'], color=sg_m['color'],
                markersize=8, capsize=4, capthick=1.3, linewidth=1.5, zorder=3)

            sig = ' *' if (row.ci_hi < 0 or row.ci_lo > 0) else ''
            ann = (f"{row.estimate:+.2f} [{row.ci_lo:+.2f}, {row.ci_hi:+.2f}]{sig}"
                   f"  n={row.n}")
            ax.text(row.ci_hi, yi, '  ' + ann,
                    va='center', fontsize=7.5, color='#333333')

        ax.axvline(0, color='#555555', linewidth=1.0, linestyle='--', alpha=0.7)
        ax.set_title(panel_title, fontweight='bold', fontsize=12)
        ax.set_xlabel(f'Change from baseline ({meta["ylabel"]})')
        ax.grid(True, alpha=0.2, axis='x')

    # Y-axis tick labels
    tp_labels_rev = list(reversed(timepoint_labels))
    tp_ticks = [y_tick_pos[t] for t in tp_labels_rev]
    axes[0].set_yticks(tp_ticks)
    axes[0].set_yticklabels(tp_labels_rev, fontsize=11, fontweight='bold')

    # Faint dividers between timepoint groups
    for ax in axes:
        for i in range(len(tp_labels_rev) - 1):
            mid_y = (y_tick_pos[tp_labels_rev[i]] + y_tick_pos[tp_labels_rev[i + 1]]) / 2
            ax.axhline(mid_y, color='#cccccc', linewidth=0.5, alpha=0.5)

    for ax in axes:
        ax.relim()
        ax.autoscale_view(scalex=True, scaley=False)

    legend_handles = [
        Line2D([0], [0], marker=sg['marker'], color=sg['color'],
               label=sg['name'], markersize=8, linestyle='None')
        for sg in subgroups
    ]
    axes[1].legend(handles=legend_handles, loc='lower right', fontsize=8, framealpha=0.9)

    fig.suptitle(f'{meta["label"]}: GEE Point Estimates Post-GLP-1 (Weight-Adjusted)',
                 fontsize=13, fontweight='bold', y=1.01)
    fig.tight_layout()
    for ext in ('png', 'pdf'):
        fig.savefig(FIGS / f"forest_{domain}.{ext}", bbox_inches='tight', dpi=300)
    plt.close(fig)
    log.info('  Saved forest_%s.png/pdf', domain)


# ── Formatted wide xlsx ────────────────────────────────────────────────────────

def _fmt_ci(est, lo, hi):
    def s(v):
        r = round(v, 2)
        if r == 0: r = abs(r)
        return f'+{r:.2f}' if r >= 0 else f'{r:.2f}'
    return f'{s(est)} ({s(lo)}, {s(hi)})'


def make_tvc_wide_xlsx(all_point_rows):
    """Generate formatted wide xlsx for weight-adjusted CFB estimates."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from collections import defaultdict
    except ImportError:
        log.warning('openpyxl not available — skipping xlsx')
        return

    df = pd.DataFrame(all_point_rows)
    if df.empty:
        return
    wt = df[df['analysis'] == 'TVC_WtAdj'].copy()

    DARK_BLUE  = '1A3A5C'
    MID_BLUE   = '4A6FA5'
    LIGHT_BLUE = 'EAF2FF'
    thin = Side(border_style='thin', color='000000')
    full = Border(left=thin, right=thin, top=thin, bottom=thin)
    no_top = Border(left=thin, right=thin,
                    top=Side(border_style=None), bottom=thin)
    no_top_no_bot = Border(left=thin, right=thin,
                           top=Side(border_style=None),
                           bottom=Side(border_style=None))

    def hfill(c): return PatternFill('solid', fgColor=c)
    def dfill(c): return PatternFill('solid', fgColor=c) if c else PatternFill(fill_type=None)
    def ca(wrap=False): return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    def la():           return Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Domain ordering
    DOMAIN_ORDER = [
        'Depression (PHQ-9, 0\u201327 scale)',
        'General Pain Intensity (0\u201310 scale)',
        'Waist Circumference (Inches)',
        'Alcohol Use (Drinks/day)',
        'Muscle Strength (MRC, 0\u20135 scale)',
    ]
    DOMAIN_BG = [LIGHT_BLUE, None, LIGHT_BLUE, None, LIGHT_BLUE]
    TP_LABELS = ['3 mo', '6 mo', '9 mo', '12 mo']

    # Build row list: (domain, subgroup, {tp: (n, est, lo, hi)})
    rows = []
    for dom in DOMAIN_ORDER:
        sub_d = wt[wt['domain'] == dom]
        if sub_d.empty:
            continue
        for sg in sub_d['subgroup'].unique():
            sub_sg = sub_d[sub_d['subgroup'] == sg]
            tp_data = {}
            for tp in TP_LABELS:
                r = sub_sg[sub_sg['timepoint'] == tp]
                if r.empty:
                    tp_data[tp] = None
                else:
                    r = r.iloc[0]
                    tp_data[tp] = (int(r['n']), r['estimate'], r['ci_lo'], r['ci_hi'])
            rows.append((dom, sg, tp_data))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TVC Weight-Adjusted CFB'
    ws.column_dimensions['A'].width = 28.0
    ws.column_dimensions['B'].width = 32.0
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 26.0
    ws.row_dimensions[1].height = 22.0
    ws.row_dimensions[2].height = 18.0

    # Header row 1
    for coord, label in [('A1', 'Outcome'), ('B1', 'Population (N, wt-adj)'),
                          ('C1', 'Estimate (95% CI) — Weight-Adjusted CFB')]:
        c = ws[coord]
        c.value, c.font = label, Font(bold=True, color='FFFFFF')
        c.fill, c.alignment = hfill(DARK_BLUE), ca(wrap=True)
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:F1')

    # Header row 2
    for col_letter, label in zip(['C', 'D', 'E', 'F'], TP_LABELS):
        c = ws[f'{col_letter}2']
        c.value, c.font = label, Font(bold=True, color='FFFFFF')
        c.fill, c.alignment = hfill(MID_BLUE), ca()

    # Group rows by domain for merging
    dom_rows_map = defaultdict(list)
    for idx, (dom, sg, _) in enumerate(rows, start=3):
        dom_rows_map[dom].append(idx)

    for er, (dom, sg, tp_data) in enumerate(rows, start=3):
        ws.row_dimensions[er].height = 34.0
        dom_idx = DOMAIN_ORDER.index(dom) if dom in DOMAIN_ORDER else 0
        bg = DOMAIN_BG[dom_idx]

        all_er = dom_rows_map[dom]
        is_first  = (er == all_er[0])
        is_last   = (er == all_er[-1])

        # Col A
        ca_cell = ws.cell(row=er, column=1)
        if is_first:
            ca_cell.value = dom
        ca_cell.fill = dfill(bg)
        ca_cell.alignment = la()
        ca_cell.border = full if is_first else (no_top if is_last else no_top_no_bot)

        # Col B
        n_val = next((v[0] for v in tp_data.values() if v is not None), None)
        cb = ws.cell(row=er, column=2)
        cb.value = f'{sg}\n(N = {n_val})' if n_val else sg
        cb.fill, cb.alignment, cb.border = dfill(bg), la(), full

        # Cols C-F
        for col_idx, tp in enumerate(TP_LABELS, start=3):
            cell = ws.cell(row=er, column=col_idx)
            cell.fill, cell.alignment, cell.border = dfill(bg), ca(), full
            val = tp_data.get(tp)
            if val is None:
                cell.value, cell.font = '\u2014', Font(bold=False)
            else:
                _, est, lo, hi = val
                sig = (lo > 0 and hi > 0) or (lo < 0 and hi < 0)
                cell.value = _fmt_ci(est, lo, hi)
                cell.font  = Font(bold=bool(sig))

    # Merge col A for multi-row domains
    for dom, er_list in dom_rows_map.items():
        if len(er_list) > 1:
            ws.merge_cells(f'A{er_list[0]}:A{er_list[-1]}')
            c = ws.cell(row=er_list[0], column=1)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    out_path = TABS / 'point_estimates_3_6_9_12mo_tvc_wt_adj_formatted_wide.xlsx'
    wb.save(out_path)
    log.info('Saved formatted xlsx -> %s', out_path)


# ── Per-subgroup model fitting ─────────────────────────────────────────────────

def fit_subgroup_tvc(df_sg, domain, meta, weight_df, coef_rows, sg_label):
    """Fit base + weight GEE for one subgroup (pre/post + CFB).

    Returns a dict with all model results needed for combined plotting.
    """
    val       = meta['val']
    spdf      = meta['spline_df']
    extra_cov = meta.get('extra_cov', [])

    # Merge concurrent weight data
    df_w = merge_weight(df_sg, weight_df)

    n_all    = df_sg.patient_id.nunique()
    n_weight = df_w.dropna(subset=['pct_weight_change']).patient_id.nunique()

    # ── Pre/post GEE ──────────────────────────────────────────────────────
    res_b, di_b, mn_b, mx_b, ref_b = fit_gee(df_w, val, spdf,
                                              include_weight=False,
                                              extra_cov=extra_cov)
    res_w, di_w, mn_w, mx_w, ref_w = fit_gee(df_w, val, spdf,
                                              include_weight=True,
                                              extra_cov=extra_cov)

    # Collect coefficient rows
    for lbl, res in [('Base', res_b), ('+Weight', res_w)]:
        if res is None:
            continue
        ci = res.conf_int()
        for param in res.params.index:
            idx = res.params.index.get_loc(param)
            coef_rows.append({
                'domain':    meta['label'],
                'subgroup':  sg_label,
                'model':     lbl,
                'parameter': param,
                'estimate':  res.params[param],
                'se':        res.bse[param],
                'pvalue':    res.pvalues[param],
                'ci_lo':     ci.iloc[idx, 0],
                'ci_hi':     ci.iloc[idx, 1],
            })

    info = {
        'sg_label': sg_label,
        'n_all': n_all, 'n_weight': n_weight,
        'df_w': df_w,
        'res_b': res_b, 'di_b': di_b, 'mn_b': mn_b, 'mx_b': mx_b, 'ref_b': ref_b,
        'res_w': res_w, 'di_w': di_w, 'mn_w': mn_w, 'mx_w': mx_w, 'ref_w': ref_w,
    }

    # ── CFB data prep + GEE ───────────────────────────────────────────────
    base_mask = df_w['days_from_baseline'].between(-BASELINE_WINDOW, BASELINE_WINDOW)
    base_df   = (df_w[base_mask].groupby('patient_id')[val].mean()
                               .rename('baseline_score').reset_index())

    demo_avail = [c for c in BASE_COV_COLS + extra_cov if c in df_w.columns]
    fu_cols    = (['patient_id', 'days_from_baseline', val, 'pct_weight_change']
                  + demo_avail)
    fu_df = (df_w[df_w['days_from_baseline'].between(FOLLOWUP_START, FOLLOWUP_END)]
             [[c for c in fu_cols if c in df_w.columns]].copy())
    fu_df = fu_df[fu_df.patient_id.isin(set(base_df.patient_id))]
    fu_df = fu_df.merge(base_df, on='patient_id', how='inner')
    fu_df['change_score'] = fu_df[val] - fu_df['baseline_score']

    info['cfb_possible'] = fu_df.patient_id.nunique() >= 15

    if info['cfb_possible']:
        n_cfb    = fu_df.patient_id.nunique()
        n_cfb_wt = fu_df.dropna(subset=['pct_weight_change']).patient_id.nunique()

        if demo_avail:
            pt_demos = (fu_df.drop_duplicates('patient_id')
                             .set_index('patient_id')[demo_avail])
            anchor = pd.DataFrame({
                'patient_id':         fu_df['patient_id'].unique(),
                'days_from_baseline': 0.0,
                'change_score':       0.0,
                'pct_weight_change':  0.0,
            }).join(pt_demos, on='patient_id')
        else:
            anchor = pd.DataFrame({
                'patient_id':         fu_df['patient_id'].unique(),
                'days_from_baseline': 0.0,
                'change_score':       0.0,
                'pct_weight_change':  0.0,
            })

        keep_cols = (['patient_id', 'days_from_baseline', 'change_score',
                      'pct_weight_change'] + demo_avail)
        gee_df = pd.concat([
            anchor[[c for c in keep_cols if c in anchor.columns]],
            fu_df  [[c for c in keep_cols if c in fu_df.columns]],
        ], ignore_index=True)

        res_cb, di_cb, mn_cb, mx_cb, ref_cb = fit_gee(
            gee_df, 'change_score', spdf, include_weight=False, extra_cov=extra_cov)
        res_cw, di_cw, mn_cw, mx_cw, ref_cw = fit_gee(
            gee_df, 'change_score', spdf, include_weight=True, extra_cov=extra_cov)

        info.update({
            'n_cfb': n_cfb, 'n_cfb_wt': n_cfb_wt,
            'gee_df': gee_df,
            'res_cb': res_cb, 'di_cb': di_cb, 'mn_cb': mn_cb, 'mx_cb': mx_cb, 'ref_cb': ref_cb,
            'res_cw': res_cw, 'di_cw': di_cw, 'mn_cw': mn_cw, 'mx_cw': mx_cw, 'ref_cw': ref_cw,
        })

    return info


# ── Combined plotting ──────────────────────────────────────────────────────────

def plot_combined_prepost(domain, meta, subgroup_results):
    """One pre/post trajectory plot per domain with all subgroups overlaid.

    Each subgroup: solid line = base model, dashed = +weight adjusted.
    Lines use the subgroup's colour; CI band only for base model.
    """
    val = meta['val']
    fig, ax = plt.subplots(figsize=(11, 5))

    for sg in subgroup_results:
        color = sg['color']
        lbl   = sg['sg_label']

        if sg['res_b'] is not None:
            m, lo, hi = predict_gee(sg['res_b'], sg['di_b'], DAYS_GRID,
                                    sg['mn_b'], sg['mx_b'], sg['ref_b'])
            smask = support_mask(sg['df_w'], val, DAYS_GRID)
            ax.fill_between(DAYS_GRID,
                            np.where(smask, lo, np.nan),
                            np.where(smask, hi, np.nan),
                            alpha=0.12, color=color)
            ax.plot(DAYS_GRID, m, '-', color=color, linewidth=2.5,
                    label=f'{lbl} (n={sg["n_all"]:,})', zorder=3)

        if sg['res_w'] is not None:
            m, lo, hi = predict_gee(sg['res_w'], sg['di_w'], DAYS_GRID,
                                    sg['mn_w'], sg['mx_w'], sg['ref_w'])
            ax.plot(DAYS_GRID, m, '--', color=color, linewidth=2.0,
                    label=f'{lbl} +wt adj (n={sg["n_weight"]:,})', zorder=3)

    ax.axvline(0, color='#CC3333', linestyle='--', linewidth=1.2, alpha=0.7)
    ax.set_ylabel(meta['ylabel'])
    ax.set_xlabel('Days from GLP-1 Initiation')
    ax.set_xlim(DAYS_GRID[0], DAYS_GRID[-1])
    ax.xaxis.set_major_locator(mticker.MultipleLocator(60))
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(xfmt))
    ax.set_title(
        f'{meta["label"]} — Base vs. Weight-Adjusted GEE Trajectory (95% CI)',
        pad=10)
    ax.legend(fontsize=8, loc='best', framealpha=0.9)
    ax.grid(True, alpha=0.2)

    return fig


def plot_combined_cfb(domain, meta, subgroup_results):
    """One CFB plot per domain — weight-adjusted GEE curves only (3 subgroups).

    Shows only the weight-adjusted model with CI bands, matching the style of
    the main CFB analysis figure.
    """
    fig, ax = plt.subplots(figsize=(11, 5))
    has_data = False

    for sg in subgroup_results:
        if not sg.get('cfb_possible', False):
            continue
        if sg.get('res_cw') is None:
            continue
        color = sg['color']
        lbl   = sg['sg_label']

        m, lo, hi = predict_gee(sg['res_cw'], sg['di_cw'], DAYS_GRID_CFB,
                                sg['mn_cw'], sg['mx_cw'], sg['ref_cw'])
        m0 = m[0] if not np.isnan(m[0]) else 0.0
        m, lo, hi = m - m0, lo - m0, hi - m0
        fu_gee = sg['gee_df'][sg['gee_df'].days_from_baseline > 0]
        smask  = support_mask(fu_gee, 'change_score', DAYS_GRID_CFB)
        smask[0] = True
        ax.fill_between(DAYS_GRID_CFB,
                         np.where(smask, lo, np.nan),
                         np.where(smask, hi, np.nan),
                         alpha=0.15, color=color)
        ax.plot(DAYS_GRID_CFB, m, '-', color=color, linewidth=2.5,
                 label=f'{lbl} (n={sg["n_cfb"]:,})', zorder=3)
        has_data = True

    if not has_data:
        plt.close(fig)
        return None

    ax.axhline(0, color='grey', linestyle='--', linewidth=0.8, alpha=0.5)
    ax.set_ylabel(f'Change from Baseline ({meta["ylabel"]})')
    ax.set_xlabel('Days from GLP-1 Initiation')
    ax.set_xlim(0, FOLLOWUP_END)
    ax.xaxis.set_major_locator(mticker.MultipleLocator(60))
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(xfmt))
    ax.set_title(
        f'{meta["label"]} — Weight-Adjusted Change from Baseline (95% CI)',
        pad=10)
    ax.legend(fontsize=9, loc='best', framealpha=0.9)
    ax.grid(True, alpha=0.2)

    return fig





# ── Main ────────────────────────────────────────────────────────────────────

def main():
    log.info("═══ Time-varying weight covariate sensitivity ═══")
    log.info("Data dir:  %s", DATA)
    log.info("Step1 CSV: %s", STEP1_PATH)
    log.info("Output:    %s", OUT)

    weight_df = load_weight_series()
    if weight_df.empty:
        log.error("No weight data loaded — cannot run time-varying analysis.")
        return

    all_point_rows = []

    for domain, meta in DOMAINS.items():
        log.info("── %s ──", meta['label'])

        df_all = load_domain(domain, meta)
        if df_all.empty:
            log.warning("  No data — skip %s", domain)
            continue

        val      = meta['val']

        # Resolve 'top25pct' sentinel to actual numeric threshold before use
        thr1 = meta.get('elev_threshold')
        if thr1 == 'top25pct':
            base_rows = df_all[df_all['days_from_baseline'].between(
                -BASELINE_WINDOW, BASELINE_WINDOW)]
            thr1 = float(base_rows.groupby('patient_id')[val].mean().quantile(0.75))
            log.info("  top25pct threshold for %s: %.2f", val, thr1)
            meta = dict(meta)  # shallow copy so we don't mutate the global
            meta['elev_label'] = f"Top 25% baseline drinkers (score \u2265{thr1:.1f})"

        df_elev  = filter_elevated(df_all, val, thr1)
        df_elev2 = filter_elevated(df_all, val, meta.get('elev_threshold_2'))

        log.info("  All: %d pts | Elev≥%s: %s | Elev≥%s: %s",
                 df_all.patient_id.nunique(),
                 thr1 if thr1 is not None else '—',
                 df_elev.patient_id.nunique()  if df_elev  is not None else 'n/a',
                 meta.get('elev_threshold_2', '—'),
                 df_elev2.patient_id.nunique() if df_elev2 is not None else 'n/a')

        subgroups = [('all', 'All patients', meta['color_all'], df_all)]
        if df_elev is not None:
            subgroups.append(('elev', meta['elev_label'], meta['color_elev'], df_elev))
        if df_elev2 is not None and meta.get('elev_threshold_2') is not None:
            subgroups.append(('elev2', meta['elev_label_2'], meta['color_elev_2'], df_elev2))

        coef_rows      = []
        subgroup_results = []

        for sg_key, sg_label, color, df_sg in subgroups:
            log.info("  Subgroup: %s (%d pts)", sg_label, df_sg.patient_id.nunique())
            info = fit_subgroup_tvc(df_sg, domain, meta, weight_df, coef_rows, sg_label)
            info['color']  = color
            info['sg_key'] = sg_key
            subgroup_results.append(info)

        # Pre/post trajectory plot
        fig_pre = plot_combined_prepost(domain, meta, subgroup_results)
        for ext in ('png', 'pdf'):
            fig_pre.savefig(FIGS / f"tvc_{domain}.{ext}")
        plt.close(fig_pre)
        log.info("  Saved tvc_%s.png/pdf", domain)

        # Weight-adjusted CFB trajectory plot
        fig_cfb = plot_combined_cfb(domain, meta, subgroup_results)
        if fig_cfb is not None:
            for ext in ('png', 'pdf'):
                fig_cfb.savefig(FIGS / f"cfb_{domain}.{ext}")
            plt.close(fig_cfb)
            log.info("  Saved cfb_%s.png/pdf", domain)

        if coef_rows:
            pd.DataFrame(coef_rows).to_csv(TABS / f"tvc_coefs_{domain}.csv",
                                           index=False)
            log.info("  Saved tvc_coefs_%s.csv", domain)

        # Point estimates at 3/6/9/12 months
        pt_rows = extract_cfb_point_estimates(meta['label'], subgroup_results)
        all_point_rows.extend(pt_rows)

        # Forest plot for this domain
        if pt_rows:
            est_df = pd.DataFrame(pt_rows)
            plot_forest_tvc(domain, meta, est_df)

    # Save combined point estimates CSV
    if all_point_rows:
        pt_df = pd.DataFrame(all_point_rows)
        pt_csv = TABS / 'point_estimates_3_6_9_12mo_tvc.csv'
        pt_df.to_csv(pt_csv, index=False)
        log.info('Saved point estimates CSV -> %s', pt_csv)

        # Formatted wide xlsx (weight-adjusted rows only)
        make_tvc_wide_xlsx(all_point_rows)

    log.info("═══ Time-varying covariate analysis complete → %s ═══", OUT)


if __name__ == '__main__':
    main()
